# -*- coding: utf-8 -*-
"""入力テンプレート(.xlsx)から空調制御試算ワークブック(7シート)を生成。
使い方: python build_workbook.py <入力.xlsx> <出力.xlsx>
"""
import sys, calendar, os, copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from model_decoder import decode

Fn="Meiryo UI"
BLUE=Font(name=Fn,color="0000FF");BLACK=Font(name=Fn,color="000000");GREEN=Font(name=Fn,color="008000")
BOLD=Font(name=Fn,bold=True);WB_=Font(name=Fn,bold=True,color="FFFFFF");TITLE=Font(name=Fn,bold=True,size=14)
RED=Font(name=Fn,color="C00000");LBL=Font(name=Fn,bold=True,color="1F4E78");GRN=Font(name=Fn,color="008000",bold=True)
RED_BOLD=Font(name=Fn,color="C00000",bold=True,size=11)
HDR=PatternFill("solid",fgColor="1F4E78");YEL=PatternFill("solid",fgColor="FFFF00")
LGRAY=PatternFill("solid",fgColor="F2F2F2");GREENF=PatternFill("solid",fgColor="E2EFDA");BLUEF=PatternFill("solid",fgColor="DDEBF7")
LBLF=PatternFill("solid",fgColor="EAF0F7");ORGF=PatternFill("solid",fgColor="FCE4D6")
thin=Side(style="thin",color="BFBFBF");BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
CEN=Alignment(horizontal="center",vertical="center");LEFT=Alignment(horizontal="left",vertical="center",wrap_text=True)
LT=Alignment(horizontal="left",vertical="top",wrap_text=True)
def st(c,font=BLACK,fill=None,align=None,fmt=None,border=True):
    c.font=font
    if fill:c.fill=fill
    if align:c.alignment=align
    if fmt:c.number_format=fmt
    if border:c.border=BORD
def coinc(kw):
    if kw=="" or kw is None:return 0.8
    if kw>=28:return 0.9
    if kw>=16:return 0.85
    return 0.7
def daysin(ym):
    try:
        y,m=str(ym).replace("-","/").split("/")[:2]; return calendar.monthrange(int(y),int(m))[1]
    except: return 30
def f(x,d=0.0):
    try: return float(x)
    except: return d

def read_input(path):
    wb=load_workbook(path,data_only=True)
    S=wb["設定"]; cfg={}
    for r in range(2,30):
        k=S.cell(r,1).value
        if k: cfg[str(k).split("(")[0].strip()]=S.cell(r,2).value
    M=wb["月次"]; mon=[]
    for r in range(2,200):
        if M.cell(r,1).value in (None,""): break
        mon.append([M.cell(r,c).value for c in range(1,10)])
    E=wb["機材"]; eq=[]; log=[]
    for r in range(2,500):
        if all(E.cell(r,c).value in (None,"") for c in range(1,6)): continue
        if E.cell(r,1).value in (None,"") and E.cell(r,3).value in (None,""): break
        loc=E.cell(r,1).value or ""; mk=E.cell(r,2).value or ""; md=E.cell(r,3).value or ""
        kh=str(E.cell(r,4).value or "").strip(); rt=E.cell(r,5).value
        ctrl = "〇" if (("〇" in kh) or ("○" in kh) or ("可" in kh and "不" not in kh) or kh.upper()=="O") else "×"
        if rt in (None,""):
            cap,flag=decode(md,mk)
            rt=cap
            if cap is not None: log.append(f"{loc}/{md}: 定格 {cap}kW を型番から推定")
            else: log.append(f"{loc}/{md}: 定格 推定不可→要確認")
        eq.append((loc,mk,md,ctrl,rt if rt not in (None,"") else ""))
    return cfg,mon,eq,log

def build(inp,out):
    cfg,mon,eq,log=read_input(inp)
    name=cfg.get("拠点名","拠点"); loc=cfg.get("所在地") or None
    pksh=f(cfg.get("空調ピーク割合"),0.20); ensh=f(cfg.get("空調電力量割合"),0.15)
    conv=f(cfg.get("conv"),0.50); cost=f(cfg.get("制御機器費"),80000)
    gyotai=cfg.get("業態") or ""
    cfg_item_price=f(cfg.get("主力商品の単価") or cfg.get("単価") or cfg.get("主力商品の単価(円)"), 10000)
    cfg_item_margin=f(cfg.get("主力商品の粗利益率") or cfg.get("粗利益率"), 0.40)
    if cfg_item_margin > 1.0: cfg_item_margin /= 100.0
    cfg_co2_factor=f(cfg.get("CO2排出係数") or cfg.get("CO2"), 0.381)
    mn=[str(x[0]) for x in mon]; dy=[daysin(x[0]) for x in mon]
    kwh=[f(x[1]) for x in mon]; pk=[f(x[2]) for x in mon]
    rate=[f(x[3])+f(x[4])+f(x[5])+f(x[6]) for x in mon]
    bu=[f(x[7]) for x in mon]; ct=[f(x[8]) for x in mon]
    n=len(mon)
    # auto report facts
    def mx(a): i=a.index(max(a)); return mn[i],max(a)
    def mxlow(a): i=a.index(min(a)); return mn[i],min(a)
    pkmax=mx(pk); pkmin=mxlow(pk); kwmax=mx(kwh); kwmin=mxlow(kwh)

    # Load and preserve Sheet 0 and Summary Report if they exist in the output file
    preserved_sheet_0 = None
    preserved_summary_report = None
    if os.path.exists(out):
        try:
            old_wb = load_workbook(out, data_only=False)
            
            # 1. Sheet 0
            sname_0 = None
            for sname in ["シート0_顧客提示用サマリー", "08_Executive_Summary"]:
                if sname in old_wb.sheetnames:
                    sname_0 = sname
                    break
            if sname_0:
                old_sheet = old_wb[sname_0]
                preserved_sheet_0 = {
                    "showGridLines": old_sheet.sheet_view.showGridLines,
                    "merged_ranges": [r.coord for r in old_sheet.merged_cells.ranges],
                    "row_heights": {r: rd.height for r, rd in old_sheet.row_dimensions.items() if rd.height is not None},
                    "col_widths": {c: cd.width for c, cd in old_sheet.column_dimensions.items() if cd.width is not None},
                    "cells": {}
                }
                for r in range(1, old_sheet.max_row + 1):
                    for c in range(1, old_sheet.max_column + 1):
                        src_cell = old_sheet.cell(row=r, column=c)
                        if src_cell.value is not None or src_cell.has_style:
                            cell_data = {
                                "value": src_cell.value,
                                "number_format": src_cell.number_format
                            }
                            if src_cell.has_style:
                                cell_data["font"] = copy.copy(src_cell.font)
                                cell_data["fill"] = copy.copy(src_cell.fill)
                                cell_data["alignment"] = copy.copy(src_cell.alignment)
                                cell_data["border"] = copy.copy(src_cell.border)
                            preserved_sheet_0["cells"][(r, c)] = cell_data
                print(f"INFO: Preserved existing Sheet 0 '{sname_0}' (cells: {len(preserved_sheet_0['cells'])})")

            # 2. Summary Report
            sname_rep = None
            for sname in ["サマリー報告", "08_Executive_Summary"]:
                if sname in old_wb.sheetnames and sname != sname_0:
                    sname_rep = sname
                    break
            if sname_rep:
                old_sheet = old_wb[sname_rep]
                preserved_summary_report = {
                    "showGridLines": old_sheet.sheet_view.showGridLines,
                    "merged_ranges": [r.coord for r in old_sheet.merged_cells.ranges],
                    "row_heights": {r: rd.height for r, rd in old_sheet.row_dimensions.items() if rd.height is not None},
                    "col_widths": {c: cd.width for c, cd in old_sheet.column_dimensions.items() if cd.width is not None},
                    "cells": {}
                }
                for r in range(1, old_sheet.max_row + 1):
                    for c in range(1, old_sheet.max_column + 1):
                        src_cell = old_sheet.cell(row=r, column=c)
                        if src_cell.value is not None or src_cell.has_style:
                            cell_data = {
                                "value": src_cell.value,
                                "number_format": src_cell.number_format
                            }
                            if src_cell.has_style:
                                cell_data["font"] = copy.copy(src_cell.font)
                                cell_data["fill"] = copy.copy(src_cell.fill)
                                cell_data["alignment"] = copy.copy(src_cell.alignment)
                                cell_data["border"] = copy.copy(src_cell.border)
                            preserved_summary_report["cells"][(r, c)] = cell_data
                print(f"INFO: Preserved existing Summary Report '{sname_rep}' (cells: {len(preserved_summary_report['cells'])})")
        except Exception as e:
            print(f"Warning: Could not read existing workbook for sheet preservation: {e}")

    wb=Workbook()
    PKSH="'前提・制御条件'!$B$5";ENSH="'前提・制御条件'!$B$6";CONV="'前提・制御条件'!$B$9"
    CAP=["'前提・制御条件'!$B$13","'前提・制御条件'!$B$14","'前提・制御条件'!$B$15"]
    # 前提
    s=wb.active; s.title="前提・制御条件"
    s["A1"]=f"{name} 電力 仕分け & 空調制御 試算"; s["A1"].font=TITLE
    s["A2"]="黄色=編集可能な前提。空調割合は要・実測検証(30分データ/夜間kWh)。"; s["A2"].font=Font(name=Fn,italic=True,color="C00000")
    s["A4"]="【1】空調比率の前提"; s["A4"].font=BOLD
    st(s.cell(5,1,"空調がピーク需要(kW)に占める割合"),BLACK,align=LEFT);st(s.cell(5,2,pksh),BLUE,YEL,CEN,"0%")
    st(s.cell(6,1,"空調が年間電力量(kWh)に占める割合"),BLACK,align=LEFT);st(s.cell(6,2,ensh),BLUE,YEL,CEN,"0%")
    s["A8"]="【2】空調制御の前提"; s["A8"].font=BOLD
    st(s.cell(9,1,"能力削減→正味エネ削減 換算係数 conv"),BLACK,align=LEFT);st(s.cell(9,2,conv),BLUE,YEL,CEN,"0%")
    s["A11"]="制御シナリオ（定格キャップ率）"; s["A11"].font=BOLD
    st(s.cell(12,1,"シナリオ"),WB_,HDR,CEN);st(s.cell(12,2,"キャップ率"),WB_,HDR,CEN);st(s.cell(12,3,""),WB_,HDR,CEN);st(s.cell(12,4,"内容"),WB_,HDR,CEN)
    for i,(nm,v,nt) in enumerate([("保守",0.20,"定格20%カット"),("標準",0.30,"定格30%カット"),("積極",0.40,"定格40%カット")]):
        r=13+i;st(s.cell(r,1,nm),BLACK,align=CEN);st(s.cell(r,2,v),BLUE,YEL,CEN,"0%");st(s.cell(r,3,""),BLACK,align=CEN);st(s.cell(r,4,nt),Font(name=Fn,size=9,color="595959"),align=LEFT)
    s["A17"]="【3】財務・環境価値の前提"; s["A17"].font=BOLD
    st(s.cell(18,1,"主力商品の単価(円)"),BLACK,align=LEFT);st(s.cell(18,2,cfg_item_price),BLUE,YEL,CEN,"#,##0")
    st(s.cell(19,1,"主力商品の粗利益率"),BLACK,align=LEFT);st(s.cell(19,2,cfg_item_margin),BLUE,YEL,CEN,"0.0%")
    st(s.cell(20,1,"CO2排出係数(kg-CO2/kWh)"),BLACK,align=LEFT);st(s.cell(20,2,cfg_co2_factor),BLUE,YEL,CEN,"0.000")
    for col,w in {"A":34,"B":12,"C":6,"D":40}.items(): s.column_dimensions[col].width=w
    # 月次
    m=wb.create_sheet("月次・電気料金"); m["A1"]="月次データ（実測：請求ベース）"; m["A1"].font=TITLE
    hd=["月","日数","使用量\n(kWh)","最大ﾃﾞﾏﾝﾄﾞ\n(kW)","日平均\n(kWh/日)","従量単価計\n(円/kWh)","基本料金\n単価","契約\n(kW)","月従量料金\n(円)","月基本料金\n(円)"]
    for j,h in enumerate(hd,1): st(m.cell(3,j,h),WB_,HDR,CEN)
    for i in range(n):
        r=4+i
        st(m.cell(r,1,mn[i]),BLACK,LGRAY if i%2 else None,CEN);m.cell(r,1).number_format="@"
        st(m.cell(r,2,dy[i]),BLACK,None,CEN,"0");st(m.cell(r,3,kwh[i]),BLUE,None,None,"#,##0");st(m.cell(r,4,pk[i]),BLUE,None,CEN,"0")
        st(m.cell(r,5,f"=C{r}/B{r}"),BLACK,None,None,"#,##0");st(m.cell(r,6,round(rate[i],2)),BLUE,None,None,"0.00");st(m.cell(r,7,bu[i]),BLUE,None,None,"#,##0");st(m.cell(r,8,ct[i]),BLUE,None,CEN,"0")
        st(m.cell(r,9,f"=C{r}*F{r}"),BLACK,None,None,"#,##0");st(m.cell(r,10,f"=G{r}*H{r}"),BLACK,None,None,"#,##0")
    tr=4+n
    st(m.cell(tr,1,"年計/最大"),WB_,HDR,CEN);st(m.cell(tr,2,f"=SUM(B4:B{tr-1})"),WB_,HDR,CEN,"0");st(m.cell(tr,3,f"=SUM(C4:C{tr-1})"),WB_,HDR,None,"#,##0");st(m.cell(tr,4,f"=MAX(D4:D{tr-1})"),WB_,HDR,CEN,"0")
    for j in (5,6,7,8): st(m.cell(tr,j,""),WB_,HDR)
    st(m.cell(tr,9,f"=SUM(I4:I{tr-1})"),WB_,HDR,None,"#,##0");st(m.cell(tr,10,f"=SUM(J4:J{tr-1})"),WB_,HDR,None,"#,##0")
    m["A"+str(tr+2)]="従量単価計=平日昼間＋燃料調整＋再エネ＋市場調整(入力の各成分の合算)。"; m["A"+str(tr+2)].font=Font(name=Fn,size=9,italic=True,color="595959")
    for j,w in enumerate([9,6,10,9,9,10,8,7,11,11],1): m.column_dimensions[get_column_letter(j)].width=w
    m.row_dimensions[3].height=30
    MTOT=f"月次・電気料金!C{tr}";MMAX=f"月次・電気料金!D{tr}";MENE=f"月次・電気料金!I{tr}";MBAS=f"月次・電気料金!J{tr}"
    # 空調・その他
    g=wb.create_sheet("空調・その他"); g["A1"]="エネルギー区分（空調 / その他）"; g["A1"].font=TITLE
    g["A2"]="空調=総使用量×空調割合(前提)。その他=残り。実測化には30分データ/夜間kWh。"; g["A2"].font=Font(name=Fn,italic=True,color="C00000")
    st(g.cell(4,1,"区分"),WB_,HDR,CEN);st(g.cell(4,2,"kWh/年"),WB_,HDR,CEN);st(g.cell(4,3,"比率"),WB_,HDR,CEN)
    st(g.cell(5,1,"総使用量(年)"),BLACK,LGRAY,LEFT);st(g.cell(5,2,f"={MTOT}"),GREEN,LGRAY,None,"#,##0");st(g.cell(5,3,"=B5/$B$5"),BLACK,LGRAY,None,"0.0%")
    st(g.cell(6,1,"空調"),BLACK,GREENF,LEFT);st(g.cell(6,2,f"=$B$5*{ENSH}"),BLACK,GREENF,None,"#,##0");st(g.cell(6,3,"=B6/$B$5"),BLACK,GREENF,None,"0.0%")
    st(g.cell(7,1,"その他"),BLACK,GREENF,LEFT);st(g.cell(7,2,f"=$B$5*(1-{ENSH})"),BLACK,GREENF,None,"#,##0");st(g.cell(7,3,"=B7/$B$5"),BLACK,GREENF,None,"0.0%")
    for col,w in {"A":16,"B":14,"C":10}.items(): g.column_dimensions[col].width=w
    # 機器台帳
    eqs=wb.create_sheet("機器台帳"); eqs["A1"]="空調 室外機 台帳（制御可否・定格）"; eqs["A1"].font=TITLE
    eqs["A2"]="〇=制御可。定格は冷房能力kW。空欄入力は型番から推定(別シート/ログ参照)。"; eqs["A2"].font=Font(name=Fn,italic=True,color="C00000")
    for j,h in enumerate(["#","設置場所","メーカー","型式","制御\n可否","定格\n(冷房kW)","制御対象\nkW"],1): st(eqs.cell(4,j,h),WB_,HDR,CEN)
    r=5
    for idx,(lc,mk,md,kh,rt) in enumerate(eq,1):
        fill=GREENF if kh=="〇" else None
        st(eqs.cell(r,1,idx),BLACK,fill,CEN,"0");st(eqs.cell(r,2,lc),BLACK,fill,LEFT);st(eqs.cell(r,3,mk),BLACK,fill,LEFT);st(eqs.cell(r,4,md),BLACK,fill,LEFT)
        st(eqs.cell(r,5,kh),BOLD,fill,CEN);st(eqs.cell(r,6,rt if rt!="" else None),BLACK,fill,CEN,"0.0")
        st(eqs.cell(r,7,(rt if kh=="〇" and rt!="" else None)),BLACK,fill,CEN,"0.0");r+=1
    tot=r
    st(eqs.cell(tot,2,"合計"),WB_,HDR,CEN)
    for j in (1,3,4): st(eqs.cell(tot,j,""),WB_,HDR)
    st(eqs.cell(tot,5,'=COUNT(F5:F'+str(r-1)+')&"台(定格有)"'),WB_,HDR,CEN);st(eqs.cell(tot,6,f"=SUM(F5:F{r-1})"),WB_,HDR,CEN,"0.0");st(eqs.cell(tot,7,f"=SUM(G5:G{r-1})"),WB_,HDR,CEN,"0.0")
    sb=tot+2
    st(eqs.cell(sb,2,"空調 総定格(冷房kW)"),BOLD,align=LEFT);st(eqs.cell(sb,6,f"=F{tot}"),GREEN,None,CEN,"0.0")
    st(eqs.cell(sb+1,2,"うち制御可能(kW)"),BOLD,align=LEFT);st(eqs.cell(sb+1,6,f"=G{tot}"),GREEN,None,CEN,"0.0")
    st(eqs.cell(sb+2,2,"制御可能 比率(容量)"),BOLD,YEL,LEFT);st(eqs.cell(sb+2,6,f"=IF(F{tot}=0,0,G{tot}/F{tot})"),BLACK,YEL,CEN,"0.0%")
    st(eqs.cell(sb+3,2,f"物理台数：{len(eq)}台（制御可{sum(1 for e in eq if e[3]=='〇')}／不可{sum(1 for e in eq if e[3]!='〇')}）"),RED,align=LEFT)
    RATIO=f"機器台帳!$F${sb+2}"
    for col,w in {"A":4,"B":18,"C":10,"D":24,"E":7,"F":10,"G":10}.items(): eqs.column_dimensions[col].width=w
    eqs.row_dimensions[4].height=28
    # シナリオ
    c4=wb.create_sheet("空調制御シナリオ"); c4["A1"]="空調のみ制御（制御可能な室外機を定格キャップ）— 削減試算"; c4["A1"].font=TITLE
    c4["A2"]="制御対象は台帳の〇のみ。基本料金は実量制前提。空調ピーク割合は前提シートで調整。"; c4["A2"].font=Font(name=Fn,italic=True,color="C00000")
    c4["A4"]="制御の基礎値"; c4["A4"].font=BOLD
    par=[("年間最大デマンド",f"={MMAX}","kW","#,##0","g"),("空調がピーク需要に占める割合",f"={PKSH}","","0%","g"),
    ("空調 ピーク寄与kW","=B5*B6","kW","#,##0","b"),("制御可能比率",f"={RATIO}","","0.0%","g"),
    ("制御対象 空調ピークkW","=B7*B8","kW","#,##0","b"),("空調 年間kWh","=空調・その他!B6","kWh","#,##0","g"),
    ("制御対象 空調kWh/年","=B10*B8","kWh","#,##0","b"),("限界従量単価(平均)",f"=AVERAGE(月次・電気料金!F4:F{tr-1})","円/kWh","0.00","g"),
    ("基本料金 平均単価",f"=AVERAGE(月次・電気料金!G4:G{tr-1})","円/kW月","#,##0","g"),("現状 年間基本料金",f"={MBAS}","円","#,##0","g"),("現状 年間従量料金",f"={MENE}","円","#,##0","g")]
    r=5
    for lab,ff,u,fmt,col in par:
        st(c4.cell(r,1,lab),BLACK,align=LEFT);st(c4.cell(r,2,ff),GREEN if col=="g" else BLACK,None,None,fmt);st(c4.cell(r,3,u),BLACK,align=CEN);r+=1
    c4["A18"]="シナリオ別 削減効果"; c4["A18"].font=BOLD
    for j,h in enumerate(["項目","保守 (20%)","標準 (30%)","積極 (40%)"],1): st(c4.cell(19,j,h),WB_,HDR,CEN)
    defs=[("定格キャップ率","={cap}","0%",0),("デマンド削減 ΔkW","=$B$9*{cap}","#,##0",0),("制御後 最大デマンド","=$B$5-$B$9*{cap}","#,##0",0),
    ("① 基本料金 削減/年","=$B$9*{cap}*$B$13*12","#,##0",1),("空調kWh 削減/年","=$B$11*{cap}*"+CONV,"#,##0",0),
    ("② 従量料金 削減/年","=$B$11*{cap}*"+CONV+"*$B$12","#,##0",1),
    ("合計 削減/年 (①+②)","=$B$9*{cap}*$B$13*12+$B$11*{cap}*"+CONV+"*$B$12","#,##0",1),
    ("対 年間電気代 削減率","=($B$9*{cap}*$B$13*12+$B$11*{cap}*"+CONV+"*$B$12)/($B$14+$B$15)","0.0%",1)]
    SCEN_ROW=26
    for ri,(lab,tmpl,fmt,hi) in enumerate(defs):
        r=20+ri;st(c4.cell(r,1,lab),BOLD if hi else BLACK,align=LEFT)
        for ci,cap in enumerate(CAP): st(c4.cell(r,2+ci,tmpl.replace("{cap}",cap)),BOLD if hi else BLACK,GREENF if hi else None,CEN,fmt)
    for col,w in {"A":30,"B":16,"C":16,"D":16}.items(): c4.column_dimensions[col].width=w
    # ROI
    units=[[lc,md,rt,coinc(rt)] for (lc,mk,md,kh,rt) in eq if kh=="〇" and rt!=""]
    units.sort(key=lambda x:-(x[2]*x[3]))
    ro=wb.create_sheet("室外機絞り込みROI"); ro["A1"]="室外機 絞り込み（ROI）"; ro["A1"].font=TITLE
    ro["A2"]="削減は『定格×稼働係数』に比例、制御費は1台固定。回収年=制御費÷削減。稼働係数(黄)を実態に。"; ro["A2"].font=Font(name=Fn,italic=True,color="C00000")
    st(ro.cell(4,1,"1台あたり制御機器費"),BLACK,align=LEFT);st(ro.cell(4,2,cost),BLUE,YEL,CEN,"#,##0");st(ro.cell(4,3,"円/台"),BLACK,align=CEN)
    st(ro.cell(5,1,"適用キャップ率"),BLACK,align=LEFT);st(ro.cell(5,2,0.30),BLUE,YEL,CEN,"0%")
    last=8+max(len(units),1)
    st(ro.cell(6,1,"制御対象 総加重"),BLACK,align=LEFT);st(ro.cell(6,2,f"=SUM(F9:F{last})"),BLACK,None,CEN,"#,##0.0")
    st(ro.cell(7,1,"適用キャップ時 合計削減/年"),BLACK,align=LEFT)
    st(ro.cell(7,2,"=$B$5*'空調制御シナリオ'!$B$9*'空調制御シナリオ'!$B$13*12+'空調制御シナリオ'!$B$11*$B$5*'前提・制御条件'!$B$9*'空調制御シナリオ'!$B$12"),GREEN,None,CEN,"#,##0")
    st(ro.cell(4,5,"▼ 回収年しきい値"),BOLD,BLUEF,LEFT);st(ro.cell(4,7,7.0),BLUE,YEL,CEN,"0.0");st(ro.cell(4,8,"年"),BLACK,BLUEF,CEN)
    st(ro.cell(5,5,"採用台数"),BOLD,BLUEF,LEFT);st(ro.cell(5,7,f'=COUNTIF(M9:M{last},"〇")'),BLACK,BLUEF,CEN,"0")
    st(ro.cell(6,5,"採用 制御kW"),BOLD,BLUEF,LEFT);st(ro.cell(6,7,f'=SUMIF(M9:M{last},"〇",D9:D{last})'),BLACK,BLUEF,CEN,"0.0")
    st(ro.cell(4,9,"採用 削減/年"),BOLD,BLUEF,LEFT);st(ro.cell(4,11,f'=SUMIF(M9:M{last},"〇",H9:H{last})'),GREEN,BLUEF,CEN,"#,##0")
    st(ro.cell(5,9,"採用 投資"),BOLD,BLUEF,LEFT);st(ro.cell(5,11,"=G5*$B$4"),BLACK,BLUEF,CEN,"#,##0")
    st(ro.cell(6,9,"採用 回収年"),BOLD,BLUEF,LEFT);st(ro.cell(6,11,"=IF(K4=0,0,K5/K4)"),BLACK,BLUEF,CEN,"0.0")
    st(ro.cell(7,9,"効果維持率"),BOLD,BLUEF,LEFT);st(ro.cell(7,11,"=IF($B$7=0,0,K4/$B$7)"),BLACK,BLUEF,CEN,"0.0%")
    for rr in range(4,8):
        for cc in range(5,12):
            if ro.cell(rr,cc).fill.fgColor.rgb in (None,"00000000"): ro.cell(rr,cc).fill=BLUEF
            ro.cell(rr,cc).border=BORD
    for j,h in enumerate(["順位","設置場所","型式","定格\nkW","稼働\n係数","加重","ΔkW","削減/年","回収\n年","累計\n削減","累計\n投資","累計\n回収","採否"],1): st(ro.cell(8,j,h),WB_,HDR,CEN)
    for i,(lc,md,kwv,cf) in enumerate(units):
        r=9+i
        st(ro.cell(r,1,i+1),BLACK,None,CEN,"0");st(ro.cell(r,2,lc),BLACK,None,LEFT);st(ro.cell(r,3,md),BLACK,None,LEFT)
        st(ro.cell(r,4,kwv),BLUE,None,CEN,"0.0");st(ro.cell(r,5,cf),BLUE,YEL,CEN,"0.00");st(ro.cell(r,6,f"=D{r}*E{r}"),BLACK,None,CEN,"0.0")
        st(ro.cell(r,7,f"=$B$5*'空調制御シナリオ'!$B$9*F{r}/$B$6"),BLACK,None,CEN,"0.0")
        st(ro.cell(r,8,f"=$B$7*F{r}/$B$6"),BLACK,None,None,"#,##0");st(ro.cell(r,9,f"=$B$4/H{r}"),BLACK,None,CEN,"0.0")
        st(ro.cell(r,10,f"=SUM($H$9:H{r})"),BLACK,None,None,"#,##0");st(ro.cell(r,11,f"=(ROW()-8)*$B$4"),BLACK,None,None,"#,##0");st(ro.cell(r,12,f"=K{r}/J{r}"),BLACK,None,CEN,"0.0")
        st(ro.cell(r,13,f'=IF(I{r}<=$G$4,"〇","×")'),BOLD,None,CEN)
    for col,w in {"A":4,"B":16,"C":22,"D":7,"E":7,"F":8,"G":7,"H":10,"I":7,"J":11,"K":11,"L":7,"M":7}.items(): ro.column_dimensions[col].width=w
    ro.row_dimensions[8].height=28
    # シート0_顧客提示用サマリー (front)
    rep=wb.create_sheet(title="シート0_顧客提示用サマリー", index=0)
    rep_report=wb.create_sheet(title="サマリー報告", index=1)
    wb.active = rep

    # Populate Sheet 0
    if preserved_sheet_0:
        rep.sheet_view.showGridLines = preserved_sheet_0["showGridLines"]
        for (r, c), cell_data in preserved_sheet_0["cells"].items():
            cell = rep.cell(row=r, column=c)
            cell.value = cell_data["value"]
            cell.number_format = cell_data["number_format"]
            if "font" in cell_data: cell.font = cell_data["font"]
            if "fill" in cell_data: cell.fill = cell_data["fill"]
            if "alignment" in cell_data: cell.alignment = cell_data["alignment"]
            if "border" in cell_data: cell.border = cell_data["border"]

        for coord in preserved_sheet_0["merged_ranges"]:
            try: rep.merge_cells(coord)
            except: pass

        for r, height in preserved_sheet_0["row_heights"].items():
            rep.row_dimensions[r].height = height
        for c, width in preserved_sheet_0["col_widths"].items():
            rep.column_dimensions[c].width = width
    else:
        rep.sheet_view.showGridLines=True
        for col, w in zip(["A","B","C","D","E","F","G","H","I","J"], [10, 18, 18, 20, 18, 18, 20, 16, 16, 16]):
            rep.column_dimensions[col].width = w
        rep.merge_cells("A1:J1")
        title_cell = rep.cell(1, 1, f"■ シート0_顧客提示用サマリー　　{name}様　空調デマンド制御 導入判断資料")
        st(title_cell, TITLE, align=Alignment(horizontal="left", vertical="center"), border=False)
        rep.row_dimensions[1].height = 30

        # Section 0
        rep.merge_cells("A3:J3")
        sec0_title = rep.cell(3, 1, "⓪ 同業他社や類似企業との効果比較（データ蓄積フェーズ用）")
        st(sec0_title, WB_, HDR, Alignment(horizontal="left", vertical="center"))
        rep.row_dimensions[3].height = 22

        rep.merge_cells("A4:J4")
        sec0_txt = rep.cell(4, 1, "【同業他社比較】現在は初期データ蓄積フェーズのため、類似企業との比較ロジックは待機中です。十分な件数のデータが蓄積され次第、自動判定が有効化されます。")
        st(sec0_txt, BLACK, LGRAY, LT)
        rep.row_dimensions[4].height = 30

        # Section 1
        rep.merge_cells("A6:J6")
        sec1_title = rep.cell(6, 1, "① 過去実績に基づく「もし導入していたら（IF）比較」")
        st(sec1_title, WB_, HDR, Alignment(horizontal="left", vertical="center"))
        rep.row_dimensions[6].height = 22

        headers = ["月", "現状維持\n(実績)[kW]", "デマンド削減\n(kW)", "もし導入していたら\n(IF)[kW]", 
                   "現状維持\n(実績)[kWh]", "使用量削減\n(kWh)", "もし導入していたら\n(IF)[kWh]", 
                   "現状維持\n(実績)[円]", "もし導入していたら\n(IF)[円]", "削減額\n(円)"]
        for j, h in enumerate(headers, 1):
            st(rep.cell(7, j, h), WB_, HDR, CEN)
        rep.row_dimensions[7].height = 28

        RATIO_cell = f"'機器台帳'!$F${sb+2}"
        for i in range(12):
            r = 8 + i
            mr = 4 + i
            st(rep.cell(r, 1, f"='月次・電気料金'!A{mr}"), BLACK, LGRAY if i%2 else None, CEN, fmt="@")
            st(rep.cell(r, 2, f"='月次・電気料金'!D{mr}"), BLACK, LGRAY if i%2 else None, CEN, fmt="#,##0")
            st(rep.cell(r, 3, f"=ROUND(B{r}*'前提・制御条件'!$B$5*0.30*{RATIO_cell}, 1)"), BLACK, LGRAY if i%2 else None, CEN, fmt="0.0")
            st(rep.cell(r, 4, f"=B{r}-C{r}"), BLACK, LGRAY if i%2 else None, CEN, fmt="#,##0")
            st(rep.cell(r, 5, f"='月次・電気料金'!C{mr}"), BLACK, LGRAY if i%2 else None, None, fmt="#,##0")
            st(rep.cell(r, 6, f"=ROUND(E{r}*'前提・制御条件'!$B$6*0.30*{RATIO_cell}*'前提・制御条件'!$B$9, 0)"), BLACK, LGRAY if i%2 else None, None, fmt="#,##0")
            st(rep.cell(r, 7, f"=E{r}-F{r}"), BLACK, LGRAY if i%2 else None, None, fmt="#,##0")
            st(rep.cell(r, 8, f"='月次・電気料金'!I{mr}+'月次・電気料金'!J{mr}"), BLACK, LGRAY if i%2 else None, None, fmt="#,##0")
            st(rep.cell(r, 9, f"=H{r}-J{r}"), BLACK, LGRAY if i%2 else None, None, fmt="#,##0")
            st(rep.cell(r, 10, f"=ROUND(C{r}*'月次・電気料金'!G{mr} + F{r}*'月次・電気料金'!F{mr}, 0)"), GRN, LGRAY if i%2 else None, None, fmt="#,##0")
            rep.row_dimensions[r].height = 20

        tr = 20
        st(rep.cell(tr, 1, "年計・最大"), WB_, HDR, CEN)
        st(rep.cell(tr, 2, "=MAX(B8:B19)"), WB_, HDR, CEN, fmt="#,##0")
        st(rep.cell(tr, 3, "=MAX(C8:C19)"), WB_, HDR, CEN, fmt="0.0")
        st(rep.cell(tr, 4, "=MAX(D8:D19)"), WB_, HDR, CEN, fmt="#,##0")
        st(rep.cell(tr, 5, "=SUM(E8:E19)"), WB_, HDR, None, fmt="#,##0")
        st(rep.cell(tr, 6, "=SUM(F8:F19)"), WB_, HDR, None, fmt="#,##0")
        st(rep.cell(tr, 7, "=SUM(G8:G19)"), WB_, HDR, None, fmt="#,##0")
        st(rep.cell(tr, 8, "=SUM(H8:H19)"), WB_, HDR, None, fmt="#,##0")
        st(rep.cell(tr, 9, "=SUM(I8:I19)"), WB_, HDR, None, fmt="#,##0")
        st(rep.cell(tr, 10, "=SUM(J8:J19)"), WB_, HDR, None, fmt="#,##0")
        rep.row_dimensions[tr].height = 22

        # Section 2
        rep.merge_cells("A22:J22")
        sec2_title = rep.cell(22, 1, "② 導入しなかった場合の今後10年間の財務損失コスト")
        st(sec2_title, WB_, HDR, Alignment(horizontal="left", vertical="center"))
        rep.row_dimensions[22].height = 22

        rep.merge_cells("A23:J23")
        sec2_txt = rep.cell(23, 1, '="過去1年間で【 ¥ " & TEXT(J20, "#,##0") & " 円 】のコストを既にドブに捨ててしまっている状態です。このまま対策なし（現状維持）を続けた場合、今後10年間で想定される財務損失合計は【 ¥ " & TEXT(J20*10, "#,##0") & " 円 】に達します。"')
        st(sec2_txt, RED_BOLD, YEL, Alignment(horizontal="center", vertical="center"))
        rep.row_dimensions[23].height = 30

        # Section 3
        rep.merge_cells("A25:J25")
        sec3_title = rep.cell(25, 1, "③ 純利益に基づく「本業の具体的商品・販売数量換算」")
        st(sec3_title, WB_, HDR, Alignment(horizontal="left", vertical="center"))
        rep.row_dimensions[25].height = 22

        rep.merge_cells("A26:J26")
        sec3_txt = rep.cell(26, 1, '="本システム導入による年間純削減額（¥ " & TEXT(J20, "#,##0") & " 円）は、御社の主力製品（単価 " & TEXT(\'前提・制御条件\'!$B$18, "#,##0") & " 円・粗利益率 " & TEXT(\'前提・制御条件\'!$B$19, "0%") & "）を新規に【 " & TEXT(IFERROR(ROUNDUP(J20/(\'前提・制御条件\'!$B$18*\'前提・制御条件\'!$B$19), 0), 0), "#,##0") & " 個 】余分に販売して稼ぎ出す本業の営業利益と全く同じ財務価値があります。"')
        st(sec3_txt, BLACK, None, LT)
        rep.row_dimensions[26].height = 40

        # Section 4
        rep.merge_cells("A28:J28")
        sec4_title = rep.cell(28, 1, "④ 脱炭素経営のための環境価値サマリー")
        st(sec4_title, WB_, HDR, Alignment(horizontal="left", vertical="center"))
        rep.row_dimensions[28].height = 22

        rep.merge_cells("A29:J29")
        sec4_txt1 = rep.cell(29, 1, '="年間 CO2排出削減量: " & TEXT(F20*\'前提・制御条件\'!$B$20, "#,##0") & " kg-CO2 / 年"')
        st(sec4_txt1, BLACK, None, LT)
        rep.row_dimensions[29].height = 18

        rep.merge_cells("A30:J30")
        sec4_txt2 = rep.cell(30, 1, '="杉の木換算本数: 約 " & TEXT(ROUND(F20*\'前提・制御条件\'!$B$20/8.8, 0), "#,##0") & " 本分 / 年（※1本あたり年間8.8kg吸収換算）"')
        st(sec4_txt2, BLACK, None, LT)
        rep.row_dimensions[30].height = 18

        # Section 5
        rep.merge_cells("A32:J32")
        sec5_title = rep.cell(32, 1, "⑤ 営業現場での顧客・経営層向け説明トークスクリプト（カンペ）")
        st(sec5_title, WB_, HDR, Alignment(horizontal="left", vertical="center"))
        rep.row_dimensions[32].height = 22

        rep.merge_cells("A33:J33")
        sec5_sub1 = rep.cell(33, 1, "【社長（経営層）への説明話法】")
        st(sec5_sub1, BOLD, BLUEF, LT)
        rep.row_dimensions[33].height = 18

        rep.merge_cells("A34:J34")
        sec5_txt1 = rep.cell(34, 1, '="社長、もし去年の時点ですでにこれを入れていれば、この高低差の分（合計" & TEXT(ROUND(J20/10000, 0), "#,##0") & "万円）の営業利益が、1ミリのリスクもなくそのまま会社に残っていたことになります。本業の製品を" & TEXT(ROUND(IFERROR(J20/(\'前提・制御条件\'!$B$18*\'前提・制御条件\'!$B$19), 0)/10000, 1), "#,##0.0") & "万個売るのと全く同じ利益が、すでに損失として消えてしまっていたということです。"')
        st(sec5_txt1, BLACK, BLUEF, LT)
        rep.row_dimensions[34].height = 45

        rep.row_dimensions[35].height = 10

        rep.merge_cells("A36:J36")
        sec5_sub2 = rep.cell(36, 1, "【現場（工場長）への説明話法】")
        st(sec5_sub2, BOLD, GREENF, LT)
        rep.row_dimensions[36].height = 18

        rep.merge_cells("A37:J37")
        sec5_txt2 = rep.cell(37, 1, "「室外機を1台ずつ数分間だけ順繰りに送風へ切り替えるローテーション制御のため、室温変化は1度未満で現場の快適性は一切変わりません。主電源を強制オフにするような無理な制御ではないため、コンプレッサーへの負荷や機器寿命への悪影響もゼロですのでご安心ください。」")
        st(sec5_txt2, BLACK, GREENF, LT)
        rep.row_dimensions[37].height = 45

        for r_idx in range(1, 39):
            for c_idx in range(1, 11):
                cell = rep.cell(row=r_idx, column=c_idx)
                if cell.alignment:
                    cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=True)
                else:
                    cell.alignment = Alignment(wrap_text=True)

    # Generate and insert 3 column charts for Peak Demand, Energy Usage, and Electric Fee comparisons.
    # Chart 1: Peak Demand (kW)
    chart_demand = BarChart()
    chart_demand.type = "col"
    chart_demand.style = 10
    chart_demand.title = "最大デマンド推移比較 (kW)"
    chart_demand.y_axis.title = "kW"
    chart_demand.x_axis.title = "月"
    chart_demand.width = 16
    chart_demand.height = 10
    
    # Add Current (Col B) and IF (Col D)
    ref_dem_curr = Reference(rep, min_col=2, min_row=7, max_row=19)
    ref_dem_if = Reference(rep, min_col=4, min_row=7, max_row=19)
    chart_demand.add_data(ref_dem_curr, titles_from_data=True)
    chart_demand.add_data(ref_dem_if, titles_from_data=True)
    chart_demand.series[0].graphicalProperties.solidFill = "0F172A" # Navy
    chart_demand.series[1].graphicalProperties.solidFill = "10B981" # Green
    
    cats = Reference(rep, min_col=1, min_row=8, max_row=19)
    chart_demand.set_categories(cats)
    chart_demand.dataLabels = DataLabelList()
    chart_demand.dataLabels.showVal = True
    rep.add_chart(chart_demand, "L3")

    # Chart 2: Energy Usage (kWh)
    chart_usage = BarChart()
    chart_usage.type = "col"
    chart_usage.style = 10
    chart_usage.title = "使用量推移比較 (kWh)"
    chart_usage.y_axis.title = "kWh"
    chart_usage.x_axis.title = "月"
    chart_usage.width = 16
    chart_usage.height = 10
    
    # Add Current (Col E) and IF (Col G)
    ref_use_curr = Reference(rep, min_col=5, min_row=7, max_row=19)
    ref_use_if = Reference(rep, min_col=7, min_row=7, max_row=19)
    chart_usage.add_data(ref_use_curr, titles_from_data=True)
    chart_usage.add_data(ref_use_if, titles_from_data=True)
    chart_usage.series[0].graphicalProperties.solidFill = "0F172A" # Navy
    chart_usage.series[1].graphicalProperties.solidFill = "10B981" # Green
    chart_usage.set_categories(cats)
    chart_usage.dataLabels = DataLabelList()
    chart_usage.dataLabels.showVal = True
    rep.add_chart(chart_usage, "L18")

    # Chart 3: Electric Fee (Yen)
    chart_fee = BarChart()
    chart_fee.type = "col"
    chart_fee.style = 10
    chart_fee.title = "電気料金推移比較 (円)"
    chart_fee.y_axis.title = "円"
    chart_fee.x_axis.title = "月"
    chart_fee.width = 16
    chart_fee.height = 10
    
    # Add Current (Col H) and IF (Col I)
    ref_fee_curr = Reference(rep, min_col=8, min_row=7, max_row=19)
    ref_fee_if = Reference(rep, min_col=9, min_row=7, max_row=19)
    chart_fee.add_data(ref_fee_curr, titles_from_data=True)
    chart_fee.add_data(ref_fee_if, titles_from_data=True)
    chart_fee.series[0].graphicalProperties.solidFill = "0F172A" # Navy
    chart_fee.series[1].graphicalProperties.solidFill = "10B981" # Green
    chart_fee.set_categories(cats)
    chart_fee.dataLabels = DataLabelList()
    chart_fee.dataLabels.showVal = True
    rep.add_chart(chart_fee, "L33")


    # Populate Summary Report (サマリー報告)
    if preserved_summary_report:
        rep_report.sheet_view.showGridLines = preserved_summary_report["showGridLines"]
        for (r, c), cell_data in preserved_summary_report["cells"].items():
            cell = rep_report.cell(row=r, column=c)
            cell.value = cell_data["value"]
            cell.number_format = cell_data["number_format"]
            if "font" in cell_data: cell.font = cell_data["font"]
            if "fill" in cell_data: cell.fill = cell_data["fill"]
            if "alignment" in cell_data: cell.alignment = cell_data["alignment"]
            if "border" in cell_data: cell.border = cell_data["border"]

        for coord in preserved_summary_report["merged_ranges"]:
            try: rep_report.merge_cells(coord)
            except: pass

        for r, height in preserved_summary_report["row_heights"].items():
            rep_report.row_dimensions[r].height = height
        for c, width in preserved_summary_report["col_widths"].items():
            rep_report.column_dimensions[c].width = width
    else:
        rep_report.sheet_view.showGridLines=False
        for col,w in {"A":20,"B":40,"C":14,"D":14}.items(): rep_report.column_dimensions[col].width=w
        title_text = f"{gyotai} {name}① 電力・空調制御 サマリー報告" if gyotai else f"{name}① 電力・空調制御 サマリー報告"
        rep_report["A1"]=title_text; rep_report["A1"].font=TITLE; rep_report.merge_cells("A1:D1")
        R=2
        if loc:
            rep_report.merge_cells(start_row=R,start_column=1,end_row=R,end_column=4);c=rep_report.cell(R,1,"所在地： "+str(loc));c.font=BOLD;R+=1
        R+=1
        def band(r,t,color="1F4E78"):
            rep_report.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4);c=rep_report.cell(r,1,t);c.font=Font(name=Fn,bold=True,size=11,color="FFFFFF");c.fill=PatternFill("solid",fgColor=color);c.alignment=Alignment(horizontal="left",vertical="center");rep_report.row_dimensions[r].height=22
        def tx(r,lab,val,fm=None):
            cl=rep_report.cell(r,1,lab);cl.font=LBL;cl.fill=LBLF;cl.alignment=LT;cl.border=BORD
            cv=rep_report.cell(r,2,val);cv.font=BLACK;cv.alignment=LT;cv.border=BORD
            if fm:cv.number_format=fm
            rep_report.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
            for cc in (3,4): rep_report.cell(r,cc).border=BORD
            rep_report.row_dimensions[r].height=15 if (isinstance(val,str) and len(val)<48) else 30
        band(R,"■ 概要");R+=1
        tx(R,"対象期間",f"{mn[0]}〜{mn[-1]}"+(f"（{gyotai}）" if gyotai else ""));R+=1
        tx(R,"年間使用量",f"={MTOT}","#,##0");R+=1
        tx(R,"年間最大デマンド",f"={MMAX}","#,##0");R+=1
        tx(R,"力率","100%");R+=1
        R+=1
        band(R,"■ 負荷の特徴（データから自動抽出）");R+=1
        tx(R,"デマンドの出方",f"通年高水準。{pkmax[0]}に山({pkmax[1]:.0f}kW)、{pkmin[0]}に谷({pkmin[1]:.0f}kW)。");R+=1
        tx(R,"使用量の山/谷",f"山={kwmax[0]} {kwmax[1]:,.0f}kWh ／ 谷={kwmin[0]} {kwmin[1]:,.0f}kWh。");R+=1
        tx(R,"負荷分解(前提)",f"空調={ensh*100:.0f}% / その他={100-ensh*100:.0f}%（前提シートで調整）。");R+=1
        R+=1
        band(R,"■ 空調設備・制御可否");R+=1
        tx(R,"室外機",f"{len(eq)}台（定格有 {sum(1 for e in eq if e[4] not in('',None))}台）。");R+=1
        tx(R,"制御可能比率(容量)",f"='機器台帳'!F{sb+2}","0.0%");R+=1
        R+=1
        band(R,"■ 空調制御による削減見込み（年間）","2E7D32");R+=1
        for j,h in enumerate(["シナリオ","保守(20%)","標準(30%)","積極(40%)"],1):
            c=rep_report.cell(R,j,h);c.font=Font(name=Fn,bold=True,color="FFFFFF");c.fill=HDR;c.alignment=CEN;c.border=BORD
        R+=1
        c=rep_report.cell(R,1,"合計削減/年");c.font=BOLD;c.alignment=LEFT;c.border=BORD
        for j,col in zip((2,3,4),("B","C","D")):
            cv=rep_report.cell(R,j,f"='空調制御シナリオ'!{col}{SCEN_ROW}");cv.font=GRN if col=="C" else BLACK;cv.alignment=CEN;cv.border=BORD;cv.number_format='"▲"#,##0'
        R+=1
        c=rep_report.cell(R,1,"対 年間電気代");c.font=BOLD;c.alignment=LEFT;c.border=BORD
        for j,col in zip((2,3,4),("B","C","D")):
            cv=rep_report.cell(R,j,f"='空調制御シナリオ'!{col}{SCEN_ROW+1}");cv.font=BLACK;cv.alignment=CEN;cv.border=BORD;cv.number_format="0.0%"
        R+=2
        band(R,"■ 所見・注意点","C55A11");R+=1
        notes=["・空調制御による基本料金(実量制)＋従量の削減見込み。詳細は各シート参照。",
        "※空調比率(ピーク/電力量)・稼働係数・convは前提値。30分データ・夜間kWhで実測化可能。",
        "※定格は冷房能力kW(電力kWではない)。空欄入力は型番から推定(推定ログ参照)。"]
        if any("推定不可" in x for x in log): notes.append("※一部の機器は型番から定格を推定できず(要・銘板確認)。")
        for val in notes:
            rep_report.merge_cells(start_row=R,start_column=1,end_row=R,end_column=4)
            c=rep_report.cell(R,1,val);c.font=RED if val.startswith("※") else BLACK;c.alignment=LT
            rep_report.row_dimensions[R].height=15 if len(val)<58 else 30;R+=1

        for r_idx in range(1, R):
            for c_idx in range(1, 5):
                cell = rep_report.cell(row=r_idx, column=c_idx)
                if cell.alignment:
                    cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=True)
                else:
                    cell.alignment = Alignment(wrap_text=True)
    # 推定ログ
    if log:
        lg=wb.create_sheet("定格推定ログ"); lg["A1"]="定格の型番推定ログ"; lg["A1"].font=TITLE
        lg["A2"]="入力で定格kWが空欄だった機器を型番から推定した記録。必要に応じ銘板で確認・上書きしてください。"; lg["A2"].font=Font(name=Fn,italic=True,color="C00000")
        for i,line in enumerate(log): lg.cell(4+i,1,("⚠ " if "不可" in line else "・ ")+line).font=Font(name=Fn,color="C00000" if "不可" in line else "000000")
        lg.column_dimensions["A"].width=80
    wb.save(out)
    print(f"OK: {out} （{len(eq)}台, 推定{len(log)}件）")

if __name__=="__main__":
    build(sys.argv[1], sys.argv[2])
