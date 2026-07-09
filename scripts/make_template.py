# -*- coding: utf-8 -*-
"""入力テンプレート(空調制御試算_入力.xlsx)を生成する。"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
Fn="Arial";BOLD=Font(name=Fn,bold=True);WB=Font(name=Fn,bold=True,color="FFFFFF")
HDR=PatternFill("solid",fgColor="1F4E78");YEL=PatternFill("solid",fgColor="FFF2CC")
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
def h(c,t):c.value=t;c.font=WB;c.fill=HDR;c.alignment=CEN
wb=Workbook()
s=wb.active; s.title="設定"
s["A1"]="設定項目";s["B1"]="値";h(s["A1"],"設定項目");h(s["B1"],"値")
rows=[("拠点名","○○拠点"),("業態(任意)","製造/物流/ホテル/スーパー 等"),("所在地(任意)",""),
("空調ピーク割合","0.20"),("空調電力量割合","0.15"),("conv(能力削減→正味エネ)","0.50"),("制御機器費(円/台)","80000")]
for i,(k,v) in enumerate(rows,2):
    s.cell(i,1,k).font=BOLD; c=s.cell(i,2,v); c.fill=YEL
s["D2"]="※業態の目安: スーパー/冷凍冷蔵主役→空調割合は小さく(ピーク5%・電力量4%程度)。製造→プロセス主役(15-20%)。物流・ホテル→空調主役(30-50%)。"
s["D2"].font=Font(name=Fn,size=9,color="C00000")
s.column_dimensions["A"].width=22;s.column_dimensions["B"].width=24;s.column_dimensions["D"].width=70
m=wb.create_sheet("月次")
mh=["月(YYYY/MM)","使用量kWh","最大デマンドkW","平日昼間単価","燃料調整単価","再エネ単価","市場調整単価","基本料金単価","契約kW"]
for j,t in enumerate(mh,1): h(m.cell(1,j),t)
m["A2"]="2025/04"
m["K1"]="※検針票の各単価をそのまま転記。無い成分は空欄(0扱い)。12ヶ月分を2行目以降に。"
m["K1"].font=Font(name=Fn,size=9,color="C00000")
for j,w in enumerate([12,11,12,10,10,9,10,11,8],1):
    m.column_dimensions[chr(64+j)].width=w
e=wb.create_sheet("機材")
eh=["設置場所","メーカー","型式","制御可否(〇/×)","定格kW(任意:空欄は型番推定)"]
for j,t in enumerate(eh,1): h(e.cell(1,j),t)
e["G1"]="※制御可否は 〇=制御可 / ×=不可。定格kWは分かれば記入、空欄なら型式から自動推定(推定フラグ付き)。"
e["G1"].font=Font(name=Fn,size=9,color="C00000")
for j,w in enumerate([18,12,24,14,24],1): e.column_dimensions[chr(64+j)].width=w
wb.save(sys.argv[1] if len(sys.argv)>1 else "空調制御試算_入力テンプレート.xlsx")
print("template saved")
